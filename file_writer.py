from openpyxl import Workbook

class file_writer:

	def __init__(self, file_path, sheet_name=None):
		self.workbook = Workbook()
		self.file_path = file_path
		self.row = 1

		if sheet_name == None:
			self.sheet = self.workbook.active
		else:
			self.sheet = self.workbook.create_sheet(title=sheet_name)

	def write_row_to_file(self, element):
		str_row = str(self.row)
		self.sheet['A'+str_row] = element.get("brand")
		self.sheet['B'+str_row] = element.get("name")
		self.sheet['C'+str_row] = element.get("label")
		self.sheet['D'+str_row] = element.get("price")

		self.row += 1

	def change_sheet(self, sheet_name):
		self.row = 1

		idx = -1
		counter = 0

		for name in self.workbook.sheetnames:
			if name == sheet_name:
				idx = counter
			counter += 1

		if idx == -1:
			self.sheet = self.workbook.create_sheet(title=sheet_name)
		else:
			self.sheet = self.workbook.active(idx)
			self.sheet = self.active()

	def cleanup(self):
		self.workbook.save(self.file_path)